from django.shortcuts import render

# Create your views here.
from django.http import HttpResponse
from openpyxl.workbook import Workbook
import openpyxl
import pandas as pd
import snowflake.connector
from rest_framework.views import APIView
from rest_framework.response import Response


class Audit(APIView):

    def post(self, request, format=None):
        name = request.data["user"]
        userpwd = request.data["password"]
        database = request.data["database"]
        account = request.data["account"]
        warehouse = request.data["warehouse"]
        schema = request.data["schema"]
        role = request.data["role"]

        def getColumnDtypes(dataTypes):

            dataList = []
            for x in dataTypes:
                if (x == 'NUMBER'):
                    dataList.append('int')
                elif (x == 'FLOAT'):
                    dataList.append('float')
                elif (x == 'bool'):
                    dataList.append('boolean')
                elif (x == 'DATE'):
                    dataList.append('DATE')
                elif (x == 'TIMESTAMP_LTZ'):
                    dataList.append('DATE')
                else:
                    dataList.append('varchar')

            return dataList

        url = snowflake.connector.connect(
            user=name,
            password=userpwd,
            account=account,
            warehouse=warehouse,
            database=database,
            schema=schema,
            role=role
        )

        cursor = url.cursor()
        script = """
    select a.TABLE_NAME,a.TABLE_SCHEMA,a.TABLE_CATALOG,b.COLUMN_NAME ,b.DATA_TYPE,b.ORDINAL_POSITION,jim.ROW_COUNT from snowflake.account_usage.tables a 
     left join INFORMATION_SCHEMA.columns b on(a.TABLE_NAME=b.TABLE_NAME)
    left join INFORMATION_SCHEMA.tables jim on (b.TABLE_NAME = jim.TABLE_NAME) ;

    """

        cursor.execute(script)
        # df= pd.DataFrame(cursor)
        # print('returned',df)
        # return df

        columns = [desc[0] for desc in cursor.description]
        data = cursor.fetchall()

        df = pd.DataFrame(list(data), columns=columns)
        dummy = getColumnDtypes(df['DATA_TYPE'].tolist())
        # df['target_data_type']= pd.dummy

        df['TARGET_DATA_TYPE'] = dummy
        df['SOURCE'] = 'snowflake'
        df['TARGET'] = 'oracle'
        # response = HttpResponse(content_type='application/ms-excel')
        # # decide file name
        # response['Content-Disposition'] = 'attachment; filename="audit_report6.xlsx"'
        return Response(df)
        # writer = pd.ExcelWriter('audit_report6.xlsx')
        # df.to_excel(writer, sheet_name='bar')
        # writer.save()



        #
        # if __name__ == '__main__':
        #     # enter your file path
        #     path = './audit_report6.xlsx'
        #
        #     # load excel file
        #     book = openpyxl.load_workbook(path)
        #
        #     # select the sheet
        #     sheet = book['bar']
        #
        #     print("Maximum rows before removing:", sheet.max_row)
        #     # sheet.max_row is the maximum number
        #     # of rows that the sheet have
        #     # delete_row() method removes rows, first parameter represents row
        #     # number and sencond parameter represents number of rows
        #     # to delete from the row number
        #     sheet.delete_rows(2, sheet.max_row - 1)
        #
        #     print("Maximum rows after removing:", sheet.max_row)
        #
        #     # save the file to the path
        #     path = 'C:/Users/Administrator/Downloads./openpy.xlsx'
        #     book.save(path)
        # #     filename = 'audit_report6.xlsx'
        # #     response = HttpResponse(
        # #         book,
        # #         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # #     )
        # #     response['Content-Disposition'] = 'attachment; filename=%s' % filename
        # #
        # #     return response
        #     return Response()
        #


